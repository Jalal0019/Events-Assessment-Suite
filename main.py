import requests
import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from bs4 import BeautifulSoup
from difflib import SequenceMatcher
from openpyxl.styles import Font
import threading
import queue
import os
import sys
import subprocess

# ---------------- Formations (shared across both tools) ----------------
FORMATIONS = {
    "1.كلية الهندسة": [152, 94, 172, 162, 397, 289, 314],
    "2.كلية العلوم": [155, 97, 175, 165, 406, 292, 317],
    "3.كلية الطب": [138, 88, 183, 182, 393, 184, 308],
    "4.كلية الهندسة الخوارزمي": [153, 95, 173, 163, 400, 290, 315],
    "5.كلية طب الكندي": [139, 89, 186, 185, 394, 187, 309],
    "6.طب الاسنان": [140, 90, 190, 188, 395, 191, 310],
    "7.كلية الصيدلة": [141, 91, 193, 192, 396, 194, 311],
    "8.كلية الطب البيطري": [142, 92, 196, 195, 398, 197, 312],
    "9.كلية التمريض": [143, 93, 200, 198, 399, 199, 313],
    "10.كلية علوم الهندسة الزراعية": [154, 831, 174, 164, 403, 291, 316],
    "11.كلية العلوم للبنات": [157, 99, 177, 167, 410, 294, 319],
    "12.كلية التربية البدنية وعلوم الرياضة": [158, 100, 178, 168, 413, 295, 320],
    "13.كلية التربية البدنية وعلوم الرياضة للبنات": [159, 101, 179, 169, 415, 296, 321],
    "14.كلية الفنون الجميلة": [161, 103, 181, 171, 417, 298, 323],
    "15.كلية التربية للعلوم الصرفة ( أبن الهيثم)": [160, 102, 180, 170, 416, 297, 322],
    "16.كلية العلوم السياسية": [202, 105, 222, 212, 404, 300, 325],
    "17.كلية القانون": [201, 104, 221, 211, 402, 299, 324],
    "18.كلية الاداب": [204, 107, 441, 214, 407, 302, 327],
    "19.كلية العلوم الاسلامية": [203, 106, 223, 213, 405, 301, 326],
    "20.كلية التربية أبن رشد للعلوم الانسانية": [208, 112, 229, 218, 412, 306, 330],
    "21.كلية التربية للبنات": [206, 111, 226, 216, 411, 304, 329],
    "22.كلية الاعلام": [435, 434, 437, 436, 439, 438, 440],
    "23.كلية اللغات": [209, 65, 230, 228, 414, 307, 331],
    "24.مركز الحاسبة الالكترونية": [239, 82, 254, 267, 689, 281, 340],
    "25.المركز الوطني الريادي لبحوث السرطان": [231, 74, 246, 245, 420, 273, 332],
    "26.مركز البحوث التربوية والنفسية": [233, 76, 248, 261, 425, 287, 334],
    "27.مركز الدراسات الاستراتيجية والدولية": [232, 75, 247, 260, 421, 274, 333],
    "28.مركز احياء التراث العلمي العربي": [235, 78, 250, 263, 432, 277, 336],
    "29.مركز بحوث ومتحف التاريخ الطبيعي": [234, 77, 249, 262, 427, 276, 335],
    "30.مركز بحوث السوق وحماية المستهلك": [237, 80, 252, 265, 431, 279, 338],
    "31.مركز دراسات المراة": [236, 79, 251, 264, 430, 278, 337],
    "32.مركز أبن سينا للتعليم الالكتروني": [240, 83, 255, 268, 426, 282, 341],
    "33.مركز التعليم المستمر": [238, 81, 253, 266, 428, 280, 339],
    "34.معهد الليزر للدراسات العليا": [242, 117, 257, 270, 422, 284, 343],
    "35.المعهد العالي للدراسات المحاسبية والمالية": [241, 116, 256, 269, 424, 283, 342],
    "36.مركز التخطيط الحضري والاقليمي للدراسات العليا": [244, 119, 259, 272, 419, 286, 345],
    "37.معهد الهندسة الوراثية للتقنيات الاحيائية": [243, 118, 258, 271, 421, 285, 344],
    "38.المركز الوطني للدراسات السكانية والديموغرافية": [1005, 1007, 1008, 1014, 1011, 1012, 1013],
    "39.كلية الادارة والاقتصاد": [156, 98, 176, 166, 408, 293, 318],
    "40.كلية التميز": [1382, 1384, 1386, 1388, 1390, 1392, 1394],
    "41.كلية الذكاء الاصطناعي": [1383, 1385, 1387, 1389, 1391, 1393, 1395],
    "42.المكتبة المركزية": [375, 564],
}


# ---------------- Shared helpers ----------------
def build_export_urls(user_input):
    urls = []
    if user_input in FORMATIONS:
        for cat_id in FORMATIONS[user_input]:
            urls.append(f"https://events.uobaghdad.edu.iq/export/categ/{cat_id}.json?pretty=yes")
    elif "/category/" in user_input:
        cat_id = user_input.split("/category/")[1].strip("/").split("/")[0]
        urls.append(f"https://events.uobaghdad.edu.iq/export/categ/{cat_id}.json?pretty=yes")
    else:
        raise ValueError("Invalid input: must be a category URL or تشكيل name.")
    return urls


def safe_get(obj, key, default=None):
    if isinstance(obj, dict):
        return obj.get(key, default)
    return default


# ============================================================
# ---------------- ANALYZER: helpers -------------------------
# ============================================================
def fetch_events(url):
    response = requests.get(url)
    response.raise_for_status()
    return response.json().get("results", [])


def fetch_event_page(event_id):
    url = f"https://events.uobaghdad.edu.iq/event/{event_id}/"
    response = requests.get(url)
    response.raise_for_status()
    return BeautifulSoup(response.text, "html.parser")


def has_contact_marker(soup):
    rows = soup.select("div.event-details > div.event-details-row")
    for row in rows:
        content = row.find("div", class_="event-details-content")
        if not content:
            continue
        if content.find("a", href=lambda h: h and h.startswith("mailto:")):
            return True
        if content.find("a", href=lambda h: h and h.startswith("tel:")):
            return True
        if content.find("i", class_=lambda c: c and "icon-mail" in c):
            return True
        if content.find("i", class_=lambda c: c and "icon-phone" in c):
            return True
        text = content.get_text(" ", strip=True)
        if text and ("@" in text or any(ch.isdigit() for ch in text)):
            return True
    return False


def has_label_marker(soup):
    span = soup.select_one("div.event-title h1 span.event-label")
    return bool(span and span.get_text(strip=True))


def has_material_marker(soup):
    return bool(soup.find("div", class_="event-details-label icon-attachment inline-attachments-icon"))


def score_event(ev):
    has_title = 1 if isinstance(ev.get("title"), str) and ev.get("title").strip() else 0
    desc = ev.get("description")
    has_description = 1 if isinstance(desc, str) and desc.strip() else 0
    start_date = ev.get("startDate", {})
    has_date = 1 if isinstance(start_date, dict) and start_date.get("date") else 0

    event_id = ev.get("id")
    has_contact = has_label = has_materials = 0
    if event_id:
        soup = fetch_event_page(event_id)
        has_contact = 1 if has_contact_marker(soup) else 0
        has_label = 1 if has_label_marker(soup) else 0
        has_materials = 1 if has_material_marker(soup) else 0

    keywords = ev.get("keywords", [])
    has_keywords = 1 if isinstance(keywords, list) and len(keywords) > 0 else 0

    fields = {
        "Title": has_title,
        "Description": has_description,
        "Date": has_date,
        "Contact": has_contact,
        "Label": has_label,
        "Keywords": has_keywords,
        "Materials": has_materials
    }
    points = sum(fields.values())
    return points, fields


def analyzer_save_results(rows, total_points, field_totals, max_points, year, category_title):
    df = pd.DataFrame(rows)
    summary = {"Event ID": "TOTALS", "Event Title": "", "Date": ""}
    summary.update(field_totals)

    percentage = (total_points / max_points * 100) if max_points else 0

    summary.update({
        "Total Points": total_points,
        "Missing": "",
        "Completeness (%)": f"{percentage:.1f}",
        "Perfect Score": f"{total_points} / {max_points}",
        "Event URL": ""
    })
    df = pd.concat([df, pd.DataFrame([summary])], ignore_index=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{category_title}_Year{year}_{timestamp}.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Scores", index=False)
        workbook = writer.book
        sheet = workbook["Scores"]
        new_row = sheet.max_row + 2
        sheet.cell(row=new_row, column=1, value="Overall Completeness (%)")
        sheet.cell(row=new_row, column=2, value=f"{percentage:.1f}%")
        sheet.cell(row=new_row, column=1).font = Font(bold=True, size=14)
        sheet.cell(row=new_row, column=2).font = Font(bold=True, size=14)

    return filename


# ============================================================
# ---------------- COMPARISON: helpers ------------------------
# ============================================================
def fetch_records(export_url, year):
    response = requests.get(export_url)
    response.raise_for_status()
    events = response.json().get("results", [])
    records = []
    for ev in events:
        date_str = ev.get("startDate", {}).get("date", "")
        if isinstance(date_str, str) and date_str.startswith(str(year)):
            if ev.get("title"):
                records.append((ev.get("id"), ev.get("title").strip()))
    return records


def load_titles_from_excel(filepath, column_index=0):
    df = pd.read_excel(filepath, header=None)
    if column_index >= df.shape[1]:
        raise ValueError("Invalid column index selected.")
    titles = df.iloc[:, column_index].dropna().astype(str).str.strip().tolist()
    return titles, df.shape[1]


def compare_save_results(indico_records, matches, points, year, formation_name, excel_titles_count):
    formation_titles_count = len(indico_records)
    completeness = round((points / formation_titles_count) * 100, 1) if formation_titles_count else 0
    relative_to_30 = round((completeness / 30) * 100, 1) if formation_titles_count else 0

    df = pd.DataFrame({
        "Event ID": [ev_id for ev_id, _ in indico_records],
        "Indico Titles": [title for _, title in indico_records],
        "Matched": [matches[ev_id][0] if ev_id in matches else "" for ev_id, _ in indico_records],
        "Similarity (%)": [matches[ev_id][1] if ev_id in matches else "" for ev_id, _ in indico_records]
    })

    summary = {
        "Event ID": "TOTALS",
        "Indico Titles": "",
        "Matched": f"{points} / {formation_titles_count} ({completeness}%)",
        "Similarity (%)": f"Relative to 30% baseline: {relative_to_30}% | Excel Titles Compared: {excel_titles_count} | Formation Titles Compared: {formation_titles_count}"
    }
    df = pd.concat([df, pd.DataFrame([summary])], ignore_index=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"title_comparison_{formation_name}_{year}_{timestamp}.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Comparison", index=False)

    return filename, completeness, relative_to_30, formation_titles_count


# ---------------- Theme (matches Student Attendance/Grades app style) ----------------
COLORS = {
    "bg":        "#F4F6F9",
    "header":    "#1B2A4A",
    "header_fg": "#FFFFFF",
    "accent":    "#2E86AB",
    "accent_dk": "#22668D",
    "card":      "#FFFFFF",
    "border":    "#E1E5EB",
    "text":      "#1F2937",
    "muted":     "#6B7280",
    "success":   "#1E8E5A",
    "warn":      "#B45309",
    "danger":    "#C0392B",
}

FONT_TITLE   = ("Segoe UI", 15, "bold")
FONT_SUB     = ("Segoe UI", 10)
FONT_LABEL   = ("Segoe UI", 10, "bold")
FONT_BODY    = ("Segoe UI", 10)
FONT_BUTTON  = ("Segoe UI", 11, "bold")
FONT_STATUS  = ("Segoe UI", 9)
FONT_TAB     = ("Segoe UI", 10, "bold")

analyzer_queue = queue.Queue()
compare_queue = queue.Queue()
analyzer_last_output = {"path": None}
compare_last_output = {"path": None}


def open_path(path):
    if not path or not os.path.exists(path):
        messagebox.showwarning("تنبيه", "لم يتم العثور على الملف.")
        return
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.call(["open", path])
        else:
            subprocess.call(["xdg-open", path])
    except Exception as e:
        messagebox.showerror("خطأ", f"تعذر فتح الملف:\n{e}")


# ============================================================
# ---------------- ANALYZER TAB LOGIC -------------------------
# ============================================================
def analyzer_worker(user_input, year):
    def progress(msg, percent=None):
        analyzer_queue.put(("progress", (msg, percent)))

    try:
        progress("بناء روابط التصدير… Building export URLs…", 0)
        export_urls = build_export_urls(user_input)
        all_rows, total_points, max_points = [], 0, 0
        field_totals = {k: 0 for k in
                         ["Title", "Description", "Date", "Contact", "Label", "Keywords", "Materials"]}

        n_urls = len(export_urls) or 1
        all_events = []
        for i, export_url in enumerate(export_urls, start=1):
            pct = int((i - 1) / n_urls * 15)
            progress(f"جلب البيانات… ({i}/{n_urls})  Fetching events…", pct)
            events = fetch_events(export_url)
            all_events.extend(events)
        progress("تم جلب جميع الفئات…", 15)

        year_events = [
            ev for ev in all_events
            if isinstance(safe_get(ev, "startDate", {}).get("date", ""), str)
            and safe_get(ev, "startDate", {}).get("date", "").startswith(str(year))
        ]

        n_events = len(year_events) or 1
        for idx, ev in enumerate(year_events, start=1):
            date_str = safe_get(ev, "startDate", {}).get("date", "")
            points, fields = score_event(ev)
            total_points += points
            max_points += 7
            for k in field_totals:
                field_totals[k] += fields[k]

            all_rows.append({
                "Event ID": ev.get("id"),
                "Event Title": ev.get("title"),
                "Date": date_str,
                **fields,
                "Total Points": points,
                "Missing": 7 - points,
                "Completeness (%)": round((points / 7) * 100, 1),
                "Perfect Score": "YES" if points == 7 else "NO",
                "Event URL": f"https://events.uobaghdad.edu.iq/event/{ev.get('id')}/"
            })

            pct = 15 + int(idx / n_events * 80)
            progress(f"تحليل النشاطات… ({idx}/{n_events})", pct)

        if all_rows:
            progress("جاري حفظ ملف الإكسل… Saving Excel file…", 97)
            filename = analyzer_save_results(all_rows, total_points, field_totals, max_points, year, user_input)
            avg_completeness = round(total_points / max_points * 100, 1) if max_points else 0
            progress("اكتمل! Done!", 100)
            analyzer_queue.put(("done", {
                "filename": filename,
                "total": total_points,
                "max": max_points,
                "avg": avg_completeness,
                "count": len(all_rows),
            }))
        else:
            analyzer_queue.put(("empty", year))
    except Exception as e:
        analyzer_queue.put(("error", str(e)))


def analyzer_start():
    user_input = analyzer_formation_combo.get().strip()
    year = analyzer_year_entry.get().strip()

    if not user_input:
        messagebox.showerror("خطأ", "الرجاء اختيار أحد التشكيلات.")
        return
    if not year.isdigit():
        messagebox.showerror("خطأ", "الرجاء إدخال سنة صحيحة، مثال: 2026")
        return

    analyzer_run_btn.config(state="disabled", text="جاري التحليل...")
    analyzer_progress["value"] = 0
    analyzer_pct_label.config(text="0%")
    analyzer_set_status("بدء التحليل… Starting analysis… (0%)", COLORS["accent"])

    thread = threading.Thread(target=analyzer_worker, args=(user_input, year), daemon=True)
    thread.start()
    root.after(100, analyzer_poll)


def analyzer_reset():
    """Reset the Analyzer tab so the user can start over with another college."""
    analyzer_formation_combo.current(0)
    analyzer_year_entry.delete(0, tk.END)
    analyzer_progress["value"] = 0
    analyzer_pct_label.config(text="0%")
    analyzer_run_btn.config(state="normal", text="اظهار النتائج على ملف اكسل")
    analyzer_open_btn.pack_forget()
    analyzer_last_output["path"] = None
    analyzer_set_status("تمت إعادة التعيين — اختر تشكيلاً جديداً", COLORS["muted"])


def analyzer_poll():
    try:
        while True:
            kind, payload = analyzer_queue.get_nowait()

            if kind == "progress":
                msg, percent = payload
                if percent is not None:
                    analyzer_progress["value"] = percent
                    analyzer_pct_label.config(text=f"{percent}%")
                    analyzer_set_status(f"{msg} ({percent}%)", COLORS["accent"])
                else:
                    analyzer_set_status(msg, COLORS["accent"])

            elif kind == "done":
                analyzer_progress["value"] = 100
                analyzer_pct_label.config(text="100%")
                analyzer_run_btn.config(state="normal", text="اظهار النتائج على ملف اكسل")
                analyzer_last_output["path"] = payload["filename"]
                analyzer_open_btn.pack(side="right", padx=(8, 0))
                analyzer_set_status(
                    f"✓ تم بنجاح — {payload['count']} نشاط — النسبة الإجمالية {payload['avg']}%",
                    COLORS["success"]
                )
                messagebox.showinfo(
                    "تم بنجاح",
                    f"تم إنشاء ملف الإكسل:\n{payload['filename']}\n\n"
                    f"عدد النقاط: {payload['total']} / {payload['max']}\n"
                    f"نسبة الاكتمال: {payload['avg']}%"
                )
                return

            elif kind == "empty":
                analyzer_progress["value"] = 0
                analyzer_pct_label.config(text="0%")
                analyzer_run_btn.config(state="normal", text="اظهار النتائج على ملف اكسل")
                analyzer_set_status(f"لا توجد نشاطات لسنة {payload}", COLORS["warn"])
                messagebox.showwarning("لا توجد بيانات", f"لم يتم العثور على أي نشاطات لسنة {payload}.")
                return

            elif kind == "error":
                analyzer_progress["value"] = 0
                analyzer_pct_label.config(text="0%")
                analyzer_run_btn.config(state="normal", text="اظهار النتائج على ملف اكسل")
                analyzer_set_status("حدث خطأ أثناء التحليل", COLORS["danger"])
                messagebox.showerror("خطأ", payload)
                return

    except queue.Empty:
        root.after(100, analyzer_poll)


def analyzer_set_status(text, color):
    analyzer_status_label.config(text=text, foreground=color)


def analyzer_open_file():
    open_path(analyzer_last_output["path"])


# ============================================================
# ---------------- COMPARISON TAB LOGIC ------------------------
# ============================================================
def compare_worker(formation_name, year, filepath, column_index):
    def progress(msg, percent=None):
        compare_queue.put(("progress", (msg, percent)))

    try:
        progress("بناء روابط التصدير… Building export URLs…", 0)
        export_urls = build_export_urls(formation_name)
        indico_records = []

        n_urls = len(export_urls) or 1
        for i, url in enumerate(export_urls, start=1):
            pct = int((i - 1) / n_urls * 50)
            progress(f"جلب عناوين الأنشطة… ({i}/{n_urls})", pct)
            indico_records.extend(fetch_records(url, int(year)))
        progress("تم جلب جميع الأنشطة…", 50)

        progress("قراءة ملف الإكسل… Reading Excel file…", 55)
        excel_titles, total_cols = load_titles_from_excel(filepath, int(column_index))

        n_records = len(indico_records) or 1
        matches = {}
        points = 0
        excel_titles_normalized = [t.strip().lower() for t in excel_titles]
        update_every = max(1, n_records // 20)

        for idx, (ev_id, title) in enumerate(indico_records, start=1):
            title_norm = title.strip().lower()
            for excel_title in excel_titles_normalized:
                similarity = SequenceMatcher(None, title_norm, excel_title).ratio()
                if similarity >= 0.8:
                    matches[ev_id] = (title, round(similarity * 100, 1))
                    points += 1
                    break
            if idx % update_every == 0 or idx == n_records:
                pct = 55 + int(idx / n_records * 40)
                progress(f"مقارنة العناوين… ({idx}/{n_records})", pct)

        progress("جاري حفظ ملف الإكسل… Saving Excel file…", 97)
        filename, completeness, relative_to_30, formation_titles_count = compare_save_results(
            indico_records, matches, points, year, formation_name, len(excel_titles)
        )
        progress("اكتمل! Done!", 100)

        compare_queue.put(("done", {
            "filename": filename,
            "points": points,
            "formation_titles_count": formation_titles_count,
            "completeness": completeness,
            "relative_to_30": relative_to_30,
            "excel_titles_count": len(excel_titles),
            "total_cols": total_cols,
            "column_index": column_index,
        }))
    except Exception as e:
        compare_queue.put(("error", str(e)))


def compare_start():
    formation_name = compare_formation_combo.get().strip()
    year = compare_year_entry.get().strip()
    filepath = compare_file_entry.get().strip()
    column_index = compare_column_entry.get().strip()

    if not formation_name:
        messagebox.showerror("خطأ", "الرجاء اختيار أحد التشكيلات.")
        return
    if not year.isdigit():
        messagebox.showerror("خطأ", "الرجاء إدخال سنة صحيحة، مثال: 2026")
        return
    if not filepath:
        messagebox.showerror("خطأ", "الرجاء اختيار ملف الإكسل.")
        return
    if not column_index.isdigit():
        messagebox.showerror("خطأ", "الرجاء إدخال رقم عمود صحيح.")
        return

    compare_run_btn.config(state="disabled", text="جاري المقارنة...")
    compare_progress["value"] = 0
    compare_pct_label.config(text="0%")
    compare_set_status("بدء المقارنة… Starting comparison… (0%)", COLORS["accent"])

    thread = threading.Thread(
        target=compare_worker,
        args=(formation_name, year, filepath, column_index),
        daemon=True
    )
    thread.start()
    root.after(100, compare_poll)


def compare_reset():
    """Reset the Comparison tab so the user can start over with another college."""
    compare_formation_combo.current(0)
    compare_year_entry.delete(0, tk.END)
    compare_file_entry.delete(0, tk.END)
    compare_column_entry.delete(0, tk.END)
    compare_progress["value"] = 0
    compare_pct_label.config(text="0%")
    compare_run_btn.config(state="normal", text="قارن Compare Titles")
    compare_open_btn.pack_forget()
    compare_last_output["path"] = None
    compare_set_status("تمت إعادة التعيين — اختر تشكيلاً جديداً", COLORS["muted"])


def compare_poll():
    try:
        while True:
            kind, payload = compare_queue.get_nowait()

            if kind == "progress":
                msg, percent = payload
                if percent is not None:
                    compare_progress["value"] = percent
                    compare_pct_label.config(text=f"{percent}%")
                    compare_set_status(f"{msg} ({percent}%)", COLORS["accent"])
                else:
                    compare_set_status(msg, COLORS["accent"])

            elif kind == "done":
                compare_progress["value"] = 100
                compare_pct_label.config(text="100%")
                compare_run_btn.config(state="normal", text="قارن Compare Titles")
                compare_last_output["path"] = payload["filename"]
                compare_open_btn.pack(side="right", padx=(8, 0))
                compare_set_status(
                    f"✓ تم بنجاح — تطابق {payload['points']} / {payload['formation_titles_count']} "
                    f"({payload['completeness']}%)",
                    COLORS["success"]
                )
                messagebox.showinfo(
                    "تم بنجاح",
                    f"تم إنشاء ملف الإكسل:\n{payload['filename']}\n\n"
                    f"إجمالي التطابقات: {payload['points']} / {payload['formation_titles_count']}\n"
                    f"نسبة الاكتمال: {payload['completeness']}%\n"
                    f"النسبة مقابل معيار 30%: {payload['relative_to_30']}%\n"
                    f"عدد عناوين الملف المقارنة: {payload['excel_titles_count']}\n"
                    f"عدد عناوين التشكيل المقارنة: {payload['formation_titles_count']}\n"
                    f"(الملف يحتوي {payload['total_cols']} عمود، تم استخدام العمود {payload['column_index']})"
                )
                return

            elif kind == "error":
                compare_progress["value"] = 0
                compare_pct_label.config(text="0%")
                compare_run_btn.config(state="normal", text="قارن Compare Titles")
                compare_set_status("حدث خطأ أثناء المقارنة", COLORS["danger"])
                messagebox.showerror("خطأ", payload)
                return

    except queue.Empty:
        root.after(100, compare_poll)


def compare_set_status(text, color):
    compare_status_label.config(text=text, foreground=color)


def compare_open_file():
    open_path(compare_last_output["path"])


def compare_browse_file():
    filepath = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if filepath:
        compare_file_entry.delete(0, tk.END)
        compare_file_entry.insert(0, filepath)


# ============================================================
# ---------------- GUI ----------------------------------------
# ============================================================
root = tk.Tk()
root.title("UoB Indico Suite  |  حزمة أدوات نشاطات إندِكو")
root.geometry("760x760")
root.configure(bg=COLORS["bg"])
root.resizable(False, False)

style = ttk.Style()
try:
    style.theme_use("clam")
except tk.TclError:
    pass

style.configure("Accent.TButton",
                font=FONT_BUTTON, padding=(14, 12),
                background=COLORS["accent"], foreground="white", borderwidth=0,
                focusthickness=0, focuscolor="none")
style.map("Accent.TButton",
          background=[("active", COLORS["accent_dk"]), ("disabled", COLORS["accent"])],
          foreground=[("disabled", "white")])

style.configure("Ghost.TButton",
                font=FONT_BODY, padding=6,
                background=COLORS["card"], foreground=COLORS["accent"], borderwidth=1)
style.map("Ghost.TButton", background=[("active", COLORS["border"])])

style.configure("Reset.TButton",
                font=FONT_BODY, padding=6,
                background=COLORS["card"], foreground=COLORS["danger"], borderwidth=1)
style.map("Reset.TButton", background=[("active", "#FBE9E7")])

style.configure("Card.TCombobox", font=FONT_BODY, padding=6)
style.configure("App.Horizontal.TProgressbar",
                troughcolor=COLORS["border"], background=COLORS["accent"],
                thickness=8, bordercolor=COLORS["border"], lightcolor=COLORS["accent"],
                darkcolor=COLORS["accent"])
style.configure("App.TEntry", padding=6)

style.configure("TNotebook", background=COLORS["bg"], borderwidth=0)
style.configure("TNotebook.Tab", font=FONT_TAB, padding=(18, 10),
                background=COLORS["border"], foreground=COLORS["text"])
style.map("TNotebook.Tab",
          background=[("selected", COLORS["card"])],
          foreground=[("selected", COLORS["accent_dk"])])

# ---- Header bar ----
header = tk.Frame(root, bg=COLORS["header"], height=72)
header.pack(fill="x", side="top")
header.pack_propagate(False)

tk.Label(header, text="جامعة بغداد  •  University of Baghdad",
         bg=COLORS["header"], fg="#9FB3D1", font=FONT_SUB).pack(anchor="w", padx=20, pady=(12, 0))
tk.Label(header, text="حزمة أدوات نشاطات Indico  —  UoB Indico Suite",
         bg=COLORS["header"], fg=COLORS["header_fg"], font=FONT_TITLE).pack(anchor="w", padx=20, pady=(0, 10))

# ---- Notebook (tabs) ----
notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True, padx=20, pady=(18, 0))

analyzer_tab = tk.Frame(notebook, bg=COLORS["bg"])
compare_tab = tk.Frame(notebook, bg=COLORS["bg"])
notebook.add(analyzer_tab, text="تقييم الأنشطة  Evaluator")
notebook.add(compare_tab, text="مقارنة العناوين  Comparison")

# ============================================================
# ---- Analyzer tab content ----
# ============================================================
analyzer_card = tk.Frame(analyzer_tab, bg=COLORS["card"], highlightbackground=COLORS["border"],
                          highlightthickness=1, bd=0)
analyzer_card.pack(fill="both", expand=True, pady=16)

analyzer_inner = tk.Frame(analyzer_card, bg=COLORS["card"])
analyzer_inner.pack(fill="both", expand=True, padx=24, pady=22)

tk.Label(analyzer_inner, text="اختر التشكيل", bg=COLORS["card"], fg=COLORS["text"],
         font=FONT_LABEL, anchor="e", justify="right").grid(row=0, column=0, columnspan=2, sticky="e", pady=(0, 4))
analyzer_formation_combo = ttk.Combobox(analyzer_inner, values=list(FORMATIONS.keys()), state="readonly",
                                         width=48, style="Card.TCombobox", justify="right")
analyzer_formation_combo.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 16))
analyzer_formation_combo.current(0)

tk.Label(analyzer_inner, text="سنة التقييم (مثال: 2026)", bg=COLORS["card"], fg=COLORS["text"],
         font=FONT_LABEL, anchor="e", justify="right").grid(row=2, column=0, columnspan=2, sticky="e", pady=(0, 4))
analyzer_year_entry = ttk.Entry(analyzer_inner, width=15, font=FONT_BODY, justify="center", style="App.TEntry")
analyzer_year_entry.grid(row=3, column=0, columnspan=2, sticky="w", pady=(0, 20))

analyzer_btn_row = tk.Frame(analyzer_inner, bg=COLORS["card"])
analyzer_btn_row.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(0, 6))

analyzer_run_btn = ttk.Button(analyzer_btn_row, text="اظهار النتائج على ملف اكسل",
                               style="Accent.TButton", command=analyzer_start, takefocus=0)
analyzer_run_btn.pack(side="right", fill="x", expand=True, ipady=6)

analyzer_open_btn = ttk.Button(analyzer_btn_row, text="فتح الملف", style="Ghost.TButton",
                                command=analyzer_open_file)
# hidden until a result exists

analyzer_reset_btn = ttk.Button(analyzer_inner, text="↺ إعادة تعيين — تشكيل جديد",
                                 style="Reset.TButton", command=analyzer_reset)
analyzer_reset_btn.grid(row=5, column=0, columnspan=2, sticky="w", pady=(4, 0))

analyzer_progress_row = tk.Frame(analyzer_inner, bg=COLORS["card"])
analyzer_progress_row.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(18, 0))
analyzer_progress_row.grid_columnconfigure(0, weight=1)

analyzer_progress = ttk.Progressbar(analyzer_progress_row, style="App.Horizontal.TProgressbar",
                                     mode="determinate", maximum=100, value=0)
analyzer_progress.grid(row=0, column=0, sticky="ew")

analyzer_pct_label = tk.Label(analyzer_progress_row, text="0%", bg=COLORS["card"],
                               fg=COLORS["accent"], font=FONT_LABEL, width=5, anchor="w")
analyzer_pct_label.grid(row=0, column=1, padx=(10, 0))

analyzer_inner.grid_columnconfigure(0, weight=1)

analyzer_footer = tk.Frame(analyzer_inner, bg=COLORS["card"])
analyzer_footer.grid(row=7, column=0, columnspan=2, sticky="ew", pady=(14, 0))
analyzer_status_label = ttk.Label(analyzer_footer,
                                   text="جاهز — كلما زاد عدد الأنشطة زاد وقت التحليل، الرجاء الانتظار",
                                   background=COLORS["card"], foreground=COLORS["muted"], font=FONT_STATUS,
                                   wraplength=620, justify="right")
analyzer_status_label.pack(anchor="e")

# ============================================================
# ---- Comparison tab content ----
# ============================================================
compare_card = tk.Frame(compare_tab, bg=COLORS["card"], highlightbackground=COLORS["border"],
                         highlightthickness=1, bd=0)
compare_card.pack(fill="both", expand=True, pady=16)

compare_inner = tk.Frame(compare_card, bg=COLORS["card"])
compare_inner.pack(fill="both", expand=True, padx=24, pady=22)

tk.Label(compare_inner, text="اختر التشكيل", bg=COLORS["card"], fg=COLORS["text"],
         font=FONT_LABEL, anchor="e", justify="right").grid(row=0, column=0, columnspan=2, sticky="e", pady=(0, 4))
compare_formation_combo = ttk.Combobox(compare_inner, values=list(FORMATIONS.keys()), state="readonly",
                                        width=48, style="Card.TCombobox", justify="right")
compare_formation_combo.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 14))
compare_formation_combo.current(0)

tk.Label(compare_inner, text="سنة التقييم (مثال: 2026)", bg=COLORS["card"], fg=COLORS["text"],
         font=FONT_LABEL, anchor="e", justify="right").grid(row=2, column=0, columnspan=2, sticky="e", pady=(0, 4))
compare_year_entry = ttk.Entry(compare_inner, width=15, font=FONT_BODY, justify="center", style="App.TEntry")
compare_year_entry.grid(row=3, column=0, columnspan=2, sticky="w", pady=(0, 14))

tk.Label(compare_inner, text="ملف الخطة (Excel)", bg=COLORS["card"], fg=COLORS["text"],
         font=FONT_LABEL, anchor="e", justify="right").grid(row=4, column=0, columnspan=2, sticky="e", pady=(0, 4))
compare_file_entry = ttk.Entry(compare_inner, width=42, font=FONT_BODY, style="App.TEntry")
compare_file_entry.grid(row=5, column=0, sticky="ew", pady=(0, 14))
ttk.Button(compare_inner, text="استعراض", style="Ghost.TButton", command=compare_browse_file).grid(
    row=5, column=1, sticky="w", padx=(8, 0), pady=(0, 14))

tk.Label(compare_inner, text="رقم العمود (0 = العمود الأول) — الافتراضي 2", bg=COLORS["card"], fg=COLORS["text"],
         font=FONT_LABEL, anchor="e", justify="right").grid(row=6, column=0, columnspan=2, sticky="e", pady=(0, 4))
compare_column_entry = ttk.Entry(compare_inner, width=15, font=FONT_BODY, justify="center", style="App.TEntry")
compare_column_entry.grid(row=7, column=0, columnspan=2, sticky="w", pady=(0, 16))

compare_btn_row = tk.Frame(compare_inner, bg=COLORS["card"])
compare_btn_row.grid(row=8, column=0, columnspan=2, sticky="ew", pady=(0, 6))

compare_run_btn = ttk.Button(compare_btn_row, text="قارن Compare Titles",
                              style="Accent.TButton", command=compare_start, takefocus=0)
compare_run_btn.pack(side="right", fill="x", expand=True, ipady=6)

compare_open_btn = ttk.Button(compare_btn_row, text="فتح الملف", style="Ghost.TButton",
                               command=compare_open_file)
# hidden until a result exists

compare_reset_btn = ttk.Button(compare_inner, text="↺ إعادة تعيين — تشكيل جديد",
                                style="Reset.TButton", command=compare_reset)
compare_reset_btn.grid(row=9, column=0, columnspan=2, sticky="w", pady=(4, 0))

compare_progress_row = tk.Frame(compare_inner, bg=COLORS["card"])
compare_progress_row.grid(row=10, column=0, columnspan=2, sticky="ew", pady=(18, 0))
compare_progress_row.grid_columnconfigure(0, weight=1)

compare_progress = ttk.Progressbar(compare_progress_row, style="App.Horizontal.TProgressbar",
                                    mode="determinate", maximum=100, value=0)
compare_progress.grid(row=0, column=0, sticky="ew")

compare_pct_label = tk.Label(compare_progress_row, text="0%", bg=COLORS["card"],
                              fg=COLORS["accent"], font=FONT_LABEL, width=5, anchor="w")
compare_pct_label.grid(row=0, column=1, padx=(10, 0))

compare_inner.grid_columnconfigure(0, weight=1)

compare_footer = tk.Frame(compare_inner, bg=COLORS["card"])
compare_footer.grid(row=11, column=0, columnspan=2, sticky="ew", pady=(14, 0))
compare_status_label = ttk.Label(compare_footer,
                                  text="جاهز — كلما زاد عدد الأنشطة زاد وقت المقارنة، الرجاء الانتظار",
                                  background=COLORS["card"], foreground=COLORS["muted"], font=FONT_STATUS,
                                  wraplength=620, justify="right")
compare_status_label.pack(anchor="e")

# ---- Global footer bar ----
footer = tk.Frame(root, bg=COLORS["header"], height=28)
footer.pack(fill="x", side="bottom")
footer.pack_propagate(False)
tk.Label(footer, text="Big Data AI Club  •  College of Artificial Intelligence, University of Baghdad",
         bg=COLORS["header"], fg="#7E93BC", font=FONT_STATUS).pack(side="right", padx=14)

root.mainloop()
